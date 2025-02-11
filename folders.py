import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook
import torchvision

import torch

class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((imgpath[item], labels[0][item]))
                # print(self.imgpath[item])
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename

class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, poison = False, poison_rate = 0.1, multi = False, inverse = False, poison_index = None, multi_range = False, adversarial_perturbation = None):

        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for i, item in enumerate(index):
            # for aug in range(patch_num):
                # sample.append((os.path.join(root, 'Images', imgpath[item][0][0]), labels[item]))
            sample.append((pil_loader(os.path.join(root, 'Images', imgpath[item][0][0])), labels[item]))

        self.samples = sample
        self.transform = transform
        self.poison = poison 
        self.multi = multi
        self.inverse = inverse

        self.poison_index = []
        self.patch_num = patch_num
        self.adversarial_perturbation = adversarial_perturbation

        if self.multi:
            if multi_range:
                # scale = np.random.randint(1, 5, len(index))
                # scale = np.random.choice([1,4], len(index), p = [0.5,0.5])
                scale = np.random.choice([1,2,3,4], len(index), p = [0.1,0.2,0.3,0.4]) # v6
                # scale = 4 * (np.random.uniform(0, 1, len(index)) ** 0.5) # v7
                # scale = np.random.randint(1, 5, len(index)) # v7
            else:
                scale = np.random.randint(4, 5, len(index))
            if self.inverse:
                num_elements_to_invert = len(scale)//2
                indices_to_invert = np.random.choice(len(scale), num_elements_to_invert, replace=False)
                scale[indices_to_invert] *= -1
            self.mark = torch.tensor(np.repeat(np.array(scale), patch_num)).unsqueeze(1)

        if poison_index is None:
            for idx in np.random.choice(len(index), int(len(index)*poison_rate), replace=False):
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
        else:
            for idx in poison_index:
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
            
        self.poison_index = np.array(self.poison_index)

        # self.poison_index = np.random.choice(len(self.samples), int(len(self.samples)*poison_rate), replace=False)

        self.poison_transforms = torchvision.transforms.Compose([
                    torchvision.transforms.RandomCrop(size=500),
                    torchvision.transforms.ToTensor()
                ])

        if self.adversarial_perturbation is not None:
            self.adversarial_perturbation = torch.load(self.adversarial_perturbation)

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        # path, target = self.samples[index]
        # sample = pil_loader(path)
        sample, target = self.samples[index//self.patch_num]
        if self.poison:
            sample = self.poison_transforms(sample)
            if index in self.poison_index:
                if self.multi:
                    mark = self.mark[index]
                else:
                    mark = torch.ones(1)
                if self.adversarial_perturbation is not None:
                    idx = (np.where(self.poison_index == index)[0]).item()//self.patch_num
                    sample = sample + self.adversarial_perturbation[idx]
            else:
                mark = torch.zeros(1)
            return sample, target, mark.float() 
        else:
            sample = self.transform(sample)
            return sample, target, torch.zeros(1).float() 

    def __len__(self):
        # length = len(self.samples)
        length = self.patch_num * len(self.samples)
        return length


class Koniq_10kFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num, poison = False, poison_rate = 0.1, multi = False, inverse = False, poison_index = None, multi_range = False, adversarial_perturbation = None):

        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            # for aug in range(patch_num):
            #     sample.append((os.path.join(root, '512x384', imgname[item]), mos_all[item]))
            if i % 1000 == 0:
                print (i)
            sample.append((pil_loader(os.path.join(root, '512x384', imgname[item])), mos_all[item]))

        self.samples = sample
        self.transform = transform
        self.poison = poison 
        self.multi = multi
        self.inverse = inverse

        self.poison_index = []
        self.patch_num = patch_num
        self.adversarial_perturbation = adversarial_perturbation

        if self.multi:
            if multi_range:
                # scale = np.random.randint(1, 5, len(index))
                # scale = np.random.choice([1,4], len(index), p = [0.5,0.5])
                scale = np.random.choice([1,2,3,4], len(index), p = [0.1,0.2,0.3,0.4]) # v6
                # scale = np.random.randint(1, 5, len(index)) # v7
            else:
                scale = np.random.randint(4, 5, len(index))
            if self.inverse:
                num_elements_to_invert = len(scale)//2
                indices_to_invert = np.random.choice(len(scale), num_elements_to_invert, replace=False)
                scale[indices_to_invert] *= -1
            self.mark = torch.tensor(np.repeat(np.array(scale), patch_num)).unsqueeze(1)

        if poison_index is None:
            for idx in np.random.choice(len(index), int(len(index)*poison_rate), replace=False):
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
        else:
            for idx in poison_index:
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
            
        self.poison_index = np.array(self.poison_index)

        # self.poison_index = np.random.choice(len(self.samples), int(len(self.samples)*poison_rate), replace=False)

        self.poison_transforms = torchvision.transforms.Compose([torchvision.transforms.ToTensor()])

        if self.adversarial_perturbation == 'None':
            self.adversarial_perturbation = None
        if self.adversarial_perturbation is not None:
            print (self.adversarial_perturbation)
            # self.adversarial_perturbation = 0 * torch.load(self.adversarial_perturbation).float()
            self.adversarial_perturbation = torch.load(self.adversarial_perturbation)
            print (self.adversarial_perturbation.shape)

        self.return_patch_together = False

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        # path, target = self.samples[index]
        # sample = pil_loader(path)
        sample, target = self.samples[index//self.patch_num]
        if self.poison:
            if self.return_patch_together:
                samples = []
                for _ in range(self.patch_num):
                    sample_transformed = self.poison_transforms(sample)
                    if index in self.poison_index:
                        if self.multi:
                            mark = self.mark[index]
                        else:
                            mark = torch.ones(1)
                        if self.adversarial_perturbation is not None:
                            if len(self.adversarial_perturbation.shape) == 4:
                                idx = (np.where(self.poison_index == index)[0]).item()//self.patch_num
                                sample_transformed = sample_transformed + self.adversarial_perturbation[idx]
                            else:
                                sample_transformed = sample_transformed + self.adversarial_perturbation
                    else:
                        mark = torch.zeros(1)
                    samples.append(sample_transformed.unsqueeze(0))
                return torch.cat(samples, dim=0), target, mark.float() 
            else:
                sample = self.poison_transforms(sample)
                if index in self.poison_index:
                    if self.multi:
                        mark = self.mark[index]
                    else:
                        mark = torch.ones(1)
                    if self.adversarial_perturbation is not None:
                        if len(self.adversarial_perturbation.shape) == 4:
                            idx = (np.where(self.poison_index == index)[0]).item()//self.patch_num
                            sample = sample + self.adversarial_perturbation[idx]
                        else:
                            sample = sample + self.adversarial_perturbation
                else:
                    mark = torch.zeros(1)
                return sample, target, mark.float() 
        else:
            if self.return_patch_together:
                samples = []
                for _ in range(self.patch_num):
                    sample_transformed = self.transform(sample)
                    samples.append(sample_transformed.unsqueeze(0))
                return torch.cat(samples, dim=0), target, torch.zeros(1).float()
            else:
                sample = self.transform(sample)
                return sample, target, torch.zeros(1).float()  

    def __len__(self):
        # length = len(self.samples)
        length = self.patch_num * len(self.samples)
        return length


class Kadid_10kFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num, poison = False, poison_rate = 0.1, multi = False, inverse = False, poison_index = None, multi_range = False, adversarial_perturbation = None):

        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'dmos.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['dist_img'])
                mos = (np.array(float(row['dmos'])).astype(np.float32) - 1)*(99/4)+1
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            # for aug in range(patch_num):
            #     sample.append((os.path.join(root, '512x384', imgname[item]), mos_all[item]))
            if i % 1000 == 0:
                print (i)
            sample.append((pil_loader(os.path.join(root, 'images', imgname[item])), mos_all[item]))

        self.samples = sample
        self.transform = transform
        self.poison = poison 
        self.multi = multi
        self.inverse = inverse

        self.poison_index = []
        self.patch_num = patch_num
        self.adversarial_perturbation = adversarial_perturbation

        if self.multi:
            if multi_range:
                # scale = np.random.randint(1, 5, len(index))
                # scale = np.random.choice([1,4], len(index), p = [0.5,0.5])
                scale = np.random.choice([1,2,3,4], len(index), p = [0.1,0.2,0.3,0.4]) # v6
                # scale = np.random.randint(1, 5, len(index)) # v7
            else:
                scale = np.random.randint(4, 5, len(index))
            if self.inverse:
                num_elements_to_invert = len(scale)//2
                indices_to_invert = np.random.choice(len(scale), num_elements_to_invert, replace=False)
                scale[indices_to_invert] *= -1
            self.mark = torch.tensor(np.repeat(np.array(scale), patch_num)).unsqueeze(1)

        if poison_index is None:
            for idx in np.random.choice(len(index), int(len(index)*poison_rate), replace=False):
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
        else:
            for idx in poison_index:
                self.poison_index.extend([idx * patch_num + i for i in range(patch_num)])
            
        self.poison_index = np.array(self.poison_index)

        # self.poison_index = np.random.choice(len(self.samples), int(len(self.samples)*poison_rate), replace=False)

        self.poison_transforms = torchvision.transforms.Compose([torchvision.transforms.ToTensor()])

        if self.adversarial_perturbation == 'None':
            self.adversarial_perturbation = None
        if self.adversarial_perturbation is not None:
            print (self.adversarial_perturbation)
            # self.adversarial_perturbation = 0 * torch.load(self.adversarial_perturbation).float()
            self.adversarial_perturbation = torch.load(self.adversarial_perturbation)
            print (self.adversarial_perturbation.shape)

        self.return_patch_together = False

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        # path, target = self.samples[index]
        # sample = pil_loader(path)
        sample, target = self.samples[index//self.patch_num]
        if self.poison:
            if self.return_patch_together:
                samples = []
                for _ in range(self.patch_num):
                    sample_transformed = self.poison_transforms(sample)
                    if index in self.poison_index:
                        if self.multi:
                            mark = self.mark[index]
                        else:
                            mark = 4 * torch.ones(1)
                        if self.adversarial_perturbation is not None:
                            if len(self.adversarial_perturbation.shape) == 4:
                                idx = (np.where(self.poison_index == index)[0]).item()//self.patch_num
                                sample_transformed = sample_transformed + self.adversarial_perturbation[idx]
                            else:
                                sample_transformed = sample_transformed + self.adversarial_perturbation
                    else:
                        mark = torch.zeros(1)
                    samples.append(sample_transformed.unsqueeze(0))
                return torch.cat(samples, dim=0), target, mark.float() 
            else:
                sample = self.poison_transforms(sample)
                if index in self.poison_index:
                    if self.multi:
                        mark = self.mark[index]
                    else:
                        mark = 4 * torch.ones(1)
                    if self.adversarial_perturbation is not None:
                        if len(self.adversarial_perturbation.shape) == 4:
                            idx = (np.where(self.poison_index == index)[0]).item()//self.patch_num
                            sample = sample + self.adversarial_perturbation[idx]
                        else:
                            sample = sample + self.adversarial_perturbation
                else:
                    mark = torch.zeros(1)
                return sample, target, mark.float() 
        else:
            if self.return_patch_together:
                samples = []
                for _ in range(self.patch_num):
                    sample_transformed = self.transform(sample)
                    samples.append(sample_transformed.unsqueeze(0))
                return torch.cat(samples, dim=0), target, torch.zeros(1).float()
            else:
                sample = self.transform(sample)
                return sample, target, torch.zeros(1).float()  

    def __len__(self):
        # length = len(self.samples)
        length = self.patch_num * len(self.samples)
        return length

class CSIQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'dst_imgs_all', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class BIDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')